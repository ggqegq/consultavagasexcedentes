# ==============================================
# CONSULTOR DE VAGAS UFF - VERSÃO STREAMLIT OTIMIZADA
# Sistema de consulta detalhada de turmas e vagas
# ==============================================

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import re
import requests
import time
from bs4 import BeautifulSoup
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# ===== CONFIGURAÇÃO DA PÁGINA =====
st.set_page_config(
    page_title="Consultor de Vagas UFF - Química",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===== ESTILOS CSS MODERNOS =====
st.markdown("""
<style>
    /* ===== TEMA GERAL ===== */
    .stApp {
        background: linear-gradient(180deg, #f0f4f8 0%, #ffffff 100%);
    }
    
    /* ===== HEADER PRINCIPAL ===== */
    .main-header-container {
        background: linear-gradient(135deg, #0f2027 0%, #203a43 50%, #2c5364 100%);
        border-radius: 16px;
        padding: 2rem;
        margin-bottom: 1.5rem;
        box-shadow: 0 10px 40px rgba(0, 0, 0, 0.15);
        text-align: center;
    }
    .main-header {
        font-size: 2.5rem;
        font-weight: 800;
        color: #ffffff;
        margin-bottom: 0.5rem;
        letter-spacing: -0.5px;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
    }
    .sub-header {
        font-size: 1.15rem;
        color: #a8d0e6;
        margin-bottom: 0.25rem;
        font-weight: 400;
    }
    .developer-credit {
        font-size: 0.95rem;
        color: #7eb8da;
        margin-top: 0.5rem;
    }
    .developer-name {
        font-weight: 700;
        color: #f0c674;
    }
    
    /* ===== CARDS DE ESTATISTICAS ===== */
    .stat-card {
        background: linear-gradient(145deg, #ffffff 0%, #f8fafc 100%);
        border-radius: 16px;
        padding: 1.5rem;
        border: 1px solid #e2e8f0;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
        transition: all 0.3s ease;
        text-align: center;
    }
    .stat-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 25px rgba(0, 0, 0, 0.1);
    }
    .stat-number {
        font-size: 2.5rem;
        font-weight: 800;
        background: linear-gradient(135deg, #1e3a5f 0%, #4a90e2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    .stat-label {
        font-size: 0.9rem;
        color: #64748b;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin-top: 0.5rem;
    }
    
    /* ===== BARRA DE PROGRESSO ===== */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #1e3a5f 0%, #3b82f6 50%, #60a5fa 100%);
        border-radius: 10px;
    }
    
    /* ===== SIDEBAR ===== */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #f8fafc 0%, #e2e8f0 100%);
    }
    section[data-testid="stSidebar"] .stMarkdown {
        color: #1e293b;
    }
    section[data-testid="stSidebar"] h1, 
    section[data-testid="stSidebar"] h2, 
    section[data-testid="stSidebar"] h3 {
        color: #1e3a5f !important;
    }
    section[data-testid="stSidebar"] label {
        color: #334155 !important;
    }
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] .stMultiSelect label,
    section[data-testid="stSidebar"] .stTextInput label {
        color: #475569 !important;
        font-weight: 500;
    }
    
    /* ===== BOTOES ===== */
    .stButton > button {
        border-radius: 12px;
        font-weight: 600;
        padding: 0.75rem 1.5rem;
        transition: all 0.3s ease;
        border: none;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #1e3a5f 0%, #3b82f6 100%);
        color: white;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(59, 130, 246, 0.3);
    }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
        color: white !important;
        border-radius: 12px;
        font-weight: 600;
        padding: 0.75rem 2rem;
        border: none;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(16, 185, 129, 0.3);
    }
    
    /* ===== TABS ===== */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background-color: #f1f5f9;
        padding: 8px;
        border-radius: 12px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        padding: 10px 20px;
        font-weight: 600;
        color: #475569;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #1e3a5f 0%, #3b82f6 100%);
        color: white !important;
    }
    
    /* ===== DIVISOR ===== */
    .custom-divider {
        height: 3px;
        background: linear-gradient(90deg, transparent, #3b82f6, #1e3a5f, #3b82f6, transparent);
        margin: 2rem 0;
        border-radius: 2px;
    }
    
    /* ===== ALERTAS ===== */
    .stAlert {
        border-radius: 12px;
        border: none;
    }
    
    /* ===== DATAFRAME ===== */
    .stDataFrame {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.05);
    }
    
    /* ===== METRICAS ===== */
    [data-testid="stMetricValue"] {
        font-size: 2rem;
        font-weight: 800;
        background: linear-gradient(135deg, #1e3a5f 0%, #3b82f6 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    [data-testid="stMetricLabel"] {
        color: #64748b;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        font-size: 0.85rem;
    }
    
    /* ===== EXPANDER ===== */
    .streamlit-expanderHeader {
        background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
        border-radius: 12px;
        font-weight: 600;
        color: #334155;
    }
    
    /* ===== CARDS INFO ===== */
    .info-card {
        background: linear-gradient(145deg, #eff6ff 0%, #dbeafe 100%);
        border-radius: 16px;
        padding: 1.5rem;
        border-left: 5px solid #3b82f6;
        margin: 1rem 0;
    }
    .warning-card {
        background: linear-gradient(145deg, #fef3c7 0%, #fde68a 100%);
        border-radius: 16px;
        padding: 1.5rem;
        border-left: 5px solid #f59e0b;
        margin: 1rem 0;
    }
    .success-card {
        background: linear-gradient(145deg, #d1fae5 0%, #a7f3d0 100%);
        border-radius: 16px;
        padding: 1.5rem;
        border-left: 5px solid #10b981;
        margin: 1rem 0;
    }
    
    /* ===== FOOTER ===== */
    .footer-container {
        background: linear-gradient(135deg, #0f2027 0%, #203a43 50%, #2c5364 100%);
        border-radius: 16px;
        padding: 1.5rem;
        margin-top: 2rem;
        text-align: center;
        box-shadow: 0 -5px 20px rgba(0, 0, 0, 0.1);
    }
    .footer-text {
        color: #a8d0e6;
        font-size: 0.95rem;
    }
    .footer-highlight {
        color: #f0c674;
        font-weight: 700;
    }
    
    /* ===== ANIMACOES ===== */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(10px); }
        to { opacity: 1; transform: translateY(0); }
    }
    .animate-fade-in {
        animation: fadeIn 0.5s ease-out;
    }
    
    /* ===== SECOES ===== */
    .section-header {
        font-size: 1.5rem;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 3px solid #3b82f6;
        display: inline-block;
    }
</style>
""", unsafe_allow_html=True)

# ===== INICIALIZAR SESSION STATE =====
if 'consultor_dados' not in st.session_state:
    st.session_state.consultor_dados = None
if 'dados_turmas' not in st.session_state:
    st.session_state.dados_turmas = None
if 'processando' not in st.session_state:
    st.session_state.processando = False
if 'resultado_disponivel' not in st.session_state:
    st.session_state.resultado_disponivel = False
if 'periodo_selecionado' not in st.session_state:
    st.session_state.periodo_selecionado = None
if 'apenas_cursos_quimica' not in st.session_state:
    st.session_state.apenas_cursos_quimica = True
if 'mostrar_outros_cursos' not in st.session_state:
    st.session_state.mostrar_outros_cursos = False

# ===== CLASSE DE CONSULTA UFF DETALHADA =====
class ConsultorQuadroHorariosUFFDetalhado:
    def __init__(self, apenas_cursos_quimica=True, mostrar_outros_cursos=False, cursos_selecionados=None):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'pt-BR,pt;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })
        
        self.base_url = "https://app.uff.br/graduacao/quadrodehorarios/"
        self.cache = {}
        self.apenas_cursos_quimica = apenas_cursos_quimica
        self.mostrar_outros_cursos = mostrar_outros_cursos
        self.cursos_selecionados = cursos_selecionados or ['Química', 'Química Industrial']
        
        # Mapeamento de cursos expandido
        self.ids_cursos = {
            'Química': '28',
            'Química Industrial': '29',
            'Engenharia Química': '27',
            'Farmácia': '15'
        }
        
        self.cores_cursos = {
            'Química': 'FFE6CC',
            'Química Industrial': 'E6F3FF',
            'Engenharia Química': 'E6FFE6',
            'Farmácia': 'FFE6FF'
        }
        
        # Códigos de cursos para filtro - inclui todos os cursos selecionados
        self.codigos_cursos_filtro = self._gerar_codigos_filtro()
    
    def _gerar_codigos_filtro(self):
        """Gera lista de códigos de curso para filtro baseado nos cursos selecionados"""
        # Retorna apenas os códigos numéricos dos cursos selecionados
        mapeamento = {
            'Química': '028',
            'Química Industrial': '029',
            'Engenharia Química': '027',
            'Farmácia': '015'
        }
        codigos = []
        for curso in self.cursos_selecionados:
            if curso in mapeamento:
                codigos.append(mapeamento[curso])
        return codigos
    
    def fazer_request(self, url, use_cache=True):
        """Faz uma requisição HTTP com cache"""
        cache_key = url
        if use_cache and cache_key in self.cache:
            if time.time() - self.cache[cache_key]['timestamp'] < 300:
                return self.cache[cache_key]['response']
        
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            
            if use_cache:
                self.cache[cache_key] = {
                    'response': response,
                    'timestamp': time.time()
                }
            
            return response
        except Exception as e:
            st.warning(f"⚠️ Erro ao acessar {url}: {e}")
            return None
    
    def construir_url_busca(self, id_curso, departamento=None, periodo='20252', codigo_disciplina=None):
        """Constrói URL de busca para o quadro de horários"""
        params = {
            'utf8': '✓',
            'q[anosemestre_eq]': periodo,
            'q[disciplina_cod_departamento_eq]': '',
            'button': '',
            'q[idturno_eq]': '',
            'q[idlocalidade_eq]': '',
            'q[vagas_turma_curso_idcurso_eq]': id_curso,
            'q[disciplina_disciplinas_curriculos_idcurriculo_eq]': '',
            'q[curso_ferias_eq]': '',
            'q[idturmamodalidade_eq]': ''
        }
        
        # Se for código de disciplina específico (3 letras + 5 números)
        if codigo_disciplina:
            params['q[disciplina_nome_or_disciplina_codigo_cont]'] = codigo_disciplina.strip().upper()
        elif departamento and departamento.strip() and departamento != 'TODOS':
            params['q[disciplina_nome_or_disciplina_codigo_cont]'] = f"{departamento.strip().upper()}00"
        else:
            params['q[disciplina_nome_or_disciplina_codigo_cont]'] = ''
        
        url_parts = [f"{key}={value}" for key, value in params.items()]
        return self.base_url + "?" + "&".join(url_parts)
    
    def extrair_links_turmas_pagina(self, html_content):
        """Extrai links para páginas detalhadas das turmas"""
        soup = BeautifulSoup(html_content, 'html.parser')
        links = []
        
        tabela = soup.find('table', class_='table')
        if tabela:
            for link in tabela.find_all('a', href=True):
                href = link['href']
                if '/turmas/' in href:
                    full_url = href if href.startswith('http') else f"https://app.uff.br{href}"
                    links.append(full_url)
        else:
            for link in soup.find_all('a', href=True):
                href = link['href']
                if '/turmas/' in href and href not in links:
                    full_url = href if href.startswith('http') else f"https://app.uff.br{href}"
                    links.append(full_url)
        
        return list(set(links))
    
    def navegar_paginas(self, url_inicial, nome_curso):
        """Navega por todas as páginas de resultados"""
        todos_links = []
        pagina_atual = 1
        
        status_placeholder = st.empty()
        
        while True:
            url_pagina = f"{url_inicial}&page={pagina_atual}" if pagina_atual > 1 else url_inicial
            status_placeholder.text(f"📄 Buscando página {pagina_atual}...")
            
            response = self.fazer_request(url_pagina)
            
            if not response:
                break
            
            soup = BeautifulSoup(response.content, 'html.parser')
            links_pagina = self.extrair_links_turmas_pagina(response.content)
            
            if not links_pagina:
                break
            
            todos_links.extend(links_pagina)
            
            pagination = soup.find('ul', class_='pagination')
            if not pagination:
                break
                
            next_disabled = pagination.find('li', class_='next disabled')
            if next_disabled:
                break
            
            pagina_atual += 1
            time.sleep(0.5)
        
        status_placeholder.empty()
        return list(set(todos_links))
    
    def extrair_horarios_turma(self, soup):
        """Extrai horários da turma"""
        try:
            secao_horarios = None
            for h in soup.find_all(['h2', 'h3', 'h4', 'h5', 'strong', 'b']):
                texto = h.get_text(strip=True).lower()
                if 'horários' in texto and 'turma' in texto:
                    secao_horarios = h
                    break
            
            if secao_horarios:
                proximo_elemento = secao_horarios.find_next(['table', 'div'])
                if proximo_elemento and proximo_elemento.name == 'table':
                    tabela_horarios = proximo_elemento
                else:
                    tabela_horarios = secao_horarios.find_next('table')
                
                if tabela_horarios:
                    horarios = []
                    dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado']
                    
                    linhas = tabela_horarios.find_all('tr')
                    if len(linhas) >= 2:
                        linha_horarios = linhas[1]
                        colunas = linha_horarios.find_all(['td', 'th'])
                        
                        for i, coluna in enumerate(colunas):
                            if i >= len(dias_semana):
                                break
                            texto = coluna.get_text(strip=True)
                            if texto and texto not in dias_semana:
                                horarios.append(f"{dias_semana[i]}: {texto}")
                    
                    return ' | '.join(horarios) if horarios else 'Não informado'
        except Exception as e:
            pass
        
        return 'Não informado'
    
    def extrair_vagas_detalhadas(self, soup, curso_origem):
        """Extrai vagas detalhadas da turma - CORRIGIDO PARA OUTROS CURSOS"""
        try:
            tabela_vagas = None
            
            for elemento in soup.find_all(['h2', 'h3', 'h4', 'h5', 'strong', 'b']):
                texto = elemento.get_text(strip=True).lower()
                if 'vagas' in texto and 'alocadas' in texto:
                    for proximo in elemento.find_next_siblings():
                        if proximo.name == 'table':
                            tabela_vagas = proximo
                            break
                    if not tabela_vagas:
                        tabela_vagas = elemento.find_next('table')
                    break
            
            if not tabela_vagas:
                for tabela in soup.find_all('table'):
                    texto_tabela = tabela.get_text(strip=True).lower()
                    if 'vagas' in texto_tabela and ('reg' in texto_tabela or 'vest' in texto_tabela):
                        tabela_vagas = tabela
                        break
            
            if not tabela_vagas:
                return []
            
            vagas_encontradas = []
            linhas = tabela_vagas.find_all('tr')
            
            for linha in linhas:
                colunas = linha.find_all(['td', 'th'])
                
                # Verificar se é linha de dados (precisa de pelo menos 4 colunas numéricas)
                if len(colunas) >= 4:
                    # A primeira coluna geralmente contém "código - nome do curso"
                    primeira_coluna = colunas[0].get_text(strip=True) if colunas else ""
                    
                    # Padrão esperado: "028 - Química" ou "029 - Química Industrial"
                    match_curso = re.match(r'^(\d{3})\s*-\s*(.+)$', primeira_coluna)
                    
                    if match_curso:
                        codigo_curso = match_curso.group(1)
                        nome_curso = match_curso.group(2).strip()
                        curso_completo = f"{codigo_curso} - {nome_curso}"
                    else:
                        # Tentar extrair código de 3 dígitos de outra forma
                        codigo_match = re.search(r'\b(\d{3})\b', primeira_coluna)
                        if codigo_match:
                            codigo_curso = codigo_match.group(1)
                            # Verificar se há nome após o código
                            resto = primeira_coluna.replace(codigo_curso, '').strip()
                            resto = re.sub(r'^[\s\-]+', '', resto).strip()
                            if resto and not resto.isdigit():
                                nome_curso = resto
                                curso_completo = f"{codigo_curso} - {nome_curso}"
                            else:
                                # Usar mapeamento de nomes conhecidos
                                nomes_conhecidos = {
                                    '028': 'Química',
                                    '029': 'Química Industrial',
                                    '027': 'Engenharia Química',
                                    '015': 'Farmácia',
                                    '025': 'Física',
                                    '020': 'Matemática',
                                    '041': 'Engenharia de Telecomunicações',
                                    '042': 'Engenharia de Produção',
                                    '043': 'Engenharia Civil',
                                    '044': 'Engenharia Mecânica',
                                    '045': 'Engenharia Elétrica',
                                }
                                nome_curso = nomes_conhecidos.get(codigo_curso, f"Curso {codigo_curso}")
                                curso_completo = f"{codigo_curso} - {nome_curso}"
                        else:
                            continue
                    
                    # Extrair números das demais colunas (vagas, inscritos, etc.)
                    numeros = []
                    for i, col in enumerate(colunas[1:], 1):  # Pular primeira coluna (curso)
                        texto_col = col.get_text(strip=True)
                        # Extrair números individuais
                        nums = re.findall(r'\b(\d+)\b', texto_col)
                        for n in nums:
                            numeros.append(int(n))
                    
                    if len(numeros) >= 4:
                        try:
                            vagas_reg = numeros[0] if len(numeros) > 0 else 0
                            vagas_vest = numeros[1] if len(numeros) > 1 else 0
                            inscritos_reg = numeros[2] if len(numeros) > 2 else 0
                            inscritos_vest = numeros[3] if len(numeros) > 3 else 0
                            
                            excedentes = 0
                            candidatos = 0
                            
                            if len(numeros) >= 6:
                                excedentes = numeros[4] if len(numeros) > 4 else 0
                                candidatos = numeros[5] if len(numeros) > 5 else 0
                            
                            # Aplicar filtros - usa comparacao exata de codigos
                            incluir_curso = False
                            
                            if self.mostrar_outros_cursos:
                                incluir_curso = True
                            elif self.apenas_cursos_quimica:
                                # Comparacao exata: codigo do curso deve estar na lista de codigos permitidos
                                codigo_padrao = codigo_curso.zfill(3)
                                if codigo_padrao in self.codigos_cursos_filtro:
                                    incluir_curso = True
                            else:
                                incluir_curso = True
                            
                            if incluir_curso:
                                if excedentes == 0 and candidatos > 0 and vagas_reg > 0:
                                    if candidatos > vagas_reg:
                                        excedentes = candidatos - vagas_reg
                                
                                vaga_info = {
                                    'curso': curso_completo,
                                    'vagas_reg': vagas_reg,
                                    'vagas_vest': vagas_vest,
                                    'inscritos_reg': inscritos_reg,
                                    'inscritos_vest': inscritos_vest,
                                    'excedentes': excedentes,
                                    'candidatos': candidatos,
                                    'vagas_disponiveis_reg': max(0, vagas_reg - inscritos_reg),
                                    'vagas_disponiveis_vest': max(0, vagas_vest - inscritos_vest),
                                    'total_vagas': vagas_reg + vagas_vest,
                                    'total_inscritos': inscritos_reg + inscritos_vest,
                                    'total_vagas_disponiveis': max(0, (vagas_reg - inscritos_reg) + (vagas_vest - inscritos_vest))
                                }
                                vagas_encontradas.append(vaga_info)
                        except Exception as e:
                            continue
            
            return vagas_encontradas
            
        except Exception as e:
            return []
    
    def extrair_dados_turma_detalhado(self, url_turma, curso_origem, periodo, departamento_busca=None):
        """Extrai dados detalhados de uma turma específica"""
        try:
            response = self.fazer_request(url_turma)
            if not response:
                return []
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            titulo = soup.find('h1')
            codigo_disciplina = ''
            nome_disciplina = ''
            turma = ''
            departamento = ''
            
            if titulo:
                texto_titulo = titulo.get_text(strip=True)
                padroes = [
                    r'Turma\s+(\S+)\s+de\s+(\S+)\s+-\s+(.+)',
                    r'(\S+)\s+-\s+(.+)\s+-\s+Turma\s+(\S+)',
                    r'(.+?)\s*-\s*Turma\s+(\S+)'
                ]
                
                for padrao in padroes:
                    match = re.search(padrao, texto_titulo)
                    if match:
                        if 'Turma' in padrao and 'de' in padrao:
                            turma = match.group(1)
                            codigo_disciplina = match.group(2)
                            nome_disciplina = match.group(3)
                        elif 'Turma' in padrao:
                            codigo_disciplina = match.group(1)
                            nome_disciplina = match.group(2)
                            turma = match.group(3) if len(match.groups()) > 2 else ''
                        break
                
                if not codigo_disciplina:
                    partes = texto_titulo.split(' - ')
                    if len(partes) >= 2:
                        primeira_parte = partes[0]
                        if 'Turma' in primeira_parte:
                            turma_match = re.search(r'Turma\s+(\S+)', primeira_parte)
                            if turma_match:
                                turma = turma_match.group(1)
                                if len(partes) > 1:
                                    segunda_parte = partes[1]
                                    if len(segunda_parte.split()) > 1:
                                        partes_codigo = segunda_parte.split()
                                        codigo_disciplina = partes_codigo[0]
                                        nome_disciplina = ' '.join(partes_codigo[1:]) if len(partes_codigo) > 1 else segunda_parte
                        
                departamento = codigo_disciplina[:3] if len(codigo_disciplina) >= 3 else ''
            
            if departamento_busca and departamento_busca != 'TODOS' and departamento != departamento_busca:
                return []
            
            horarios = self.extrair_horarios_turma(soup)
            vagas_detalhadas = self.extrair_vagas_detalhadas(soup, curso_origem)
            
            if not vagas_detalhadas:
                if self.apenas_cursos_quimica and not self.mostrar_outros_cursos:
                    return []
                
                registro_basico = {
                    'periodo': periodo,
                    'departamento': departamento,
                    'codigo_disciplina': codigo_disciplina,
                    'nome_disciplina': nome_disciplina,
                    'turma': turma,
                    'horarios': horarios,
                    'curso_origem_busca': curso_origem,
                    'curso_vaga': curso_origem,
                    'vagas_reg': 0,
                    'vagas_vest': 0,
                    'inscritos_reg': 0,
                    'inscritos_vest': 0,
                    'excedentes': 0,
                    'candidatos': 0,
                    'vagas_disponiveis_reg': 0,
                    'vagas_disponiveis_vest': 0,
                    'total_vagas': 0,
                    'total_inscritos': 0,
                    'total_vagas_disponiveis': 0,
                    'url': url_turma
                }
                return [registro_basico]
            
            registros = []
            for vaga in vagas_detalhadas:
                registro = {
                    'periodo': periodo,
                    'departamento': departamento,
                    'codigo_disciplina': codigo_disciplina,
                    'nome_disciplina': nome_disciplina,
                    'turma': turma,
                    'horarios': horarios,
                    'curso_origem_busca': curso_origem,
                    'curso_vaga': vaga['curso'],
                    'vagas_reg': vaga['vagas_reg'],
                    'vagas_vest': vaga['vagas_vest'],
                    'inscritos_reg': vaga['inscritos_reg'],
                    'inscritos_vest': vaga['inscritos_vest'],
                    'excedentes': vaga['excedentes'],
                    'candidatos': vaga['candidatos'],
                    'vagas_disponiveis_reg': vaga['vagas_disponiveis_reg'],
                    'vagas_disponiveis_vest': vaga['vagas_disponiveis_vest'],
                    'total_vagas': vaga['total_vagas'],
                    'total_inscritos': vaga['total_inscritos'],
                    'total_vagas_disponiveis': vaga['total_vagas_disponiveis'],
                    'url': url_turma
                }
                registros.append(registro)
            
            return registros
            
        except Exception as e:
            return []
    
    def buscar_turmas_detalhadas(self, curso_nome, periodo, departamento=None, codigo_disciplina=None):
        """Busca turmas detalhadas com todos os dados"""
        msg = f"🔍 Buscando turmas de {curso_nome} - Período {periodo}"
        if codigo_disciplina:
            msg += f" - Disciplina {codigo_disciplina}"
        elif departamento and departamento != 'TODOS':
            msg += f" - Depto {departamento}"
        st.info(msg)
        
        id_curso = self.ids_cursos.get(curso_nome)
        if not id_curso:
            return []
        
        url_busca = self.construir_url_busca(id_curso, departamento, periodo, codigo_disciplina)
        links_turmas = self.navegar_paginas(url_busca, curso_nome)
        
        if not links_turmas:
            st.warning(f"ℹ️ Nenhuma turma encontrada para {curso_nome} no período {periodo}")
            return []
        
        todas_turmas = []
        total_turmas = len(links_turmas)
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, link in enumerate(links_turmas):
            if st.session_state.processando == False:
                break
                
            status_text.text(f"📋 Processando turma {i+1}/{total_turmas}")
            
            registros = self.extrair_dados_turma_detalhado(link, curso_nome, periodo, departamento)
            
            for registro in registros:
                duplicado = False
                for existente in todas_turmas:
                    if (existente['codigo_disciplina'] == registro['codigo_disciplina'] and
                        existente['turma'] == registro['turma'] and
                        existente['curso_vaga'] == registro['curso_vaga']):
                        duplicado = True
                        break
                
                if not duplicado:
                    todas_turmas.append(registro)
            
            progress_bar.progress((i + 1) / total_turmas)
            time.sleep(0.3)
        
        progress_bar.empty()
        status_text.empty()
        
        return todas_turmas
    
    def consultar_vagas_completas(self, periodos, cursos, departamentos, codigo_disciplina=None):
        """Consulta completa de vagas com todos os detalhes"""
        todas_turmas = []
        
        total_consultas = len(periodos) * len(cursos) * len(departamentos)
        consulta_atual = 0
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for periodo in periodos:
            for curso in cursos:
                for depto in departamentos:
                    if st.session_state.processando == False:
                        return todas_turmas
                    
                    consulta_atual += 1
                    progresso = consulta_atual / total_consultas
                    progress_bar.progress(progresso)
                    
                    status_text.text(f"🔍 {curso} | 📅 {periodo} | 🏫 {depto or 'Todos'}")
                    
                    turmas = self.buscar_turmas_detalhadas(curso, periodo, depto, codigo_disciplina)
                    
                    for turma in turmas:
                        duplicado = False
                        for existente in todas_turmas:
                            if (existente['codigo_disciplina'] == turma['codigo_disciplina'] and
                                existente['turma'] == turma['turma'] and
                                existente['curso_vaga'] == turma['curso_vaga'] and
                                existente['periodo'] == turma['periodo']):
                                duplicado = True
                                break
                        
                        if not duplicado:
                            todas_turmas.append(turma)
                    
                    time.sleep(0.5)
        
        progress_bar.empty()
        status_text.empty()
        
        return todas_turmas

# ===== FUNÇÕES PARA FORMATAÇÃO EXCEL =====
def aplicar_formatacao_excel(workbook):
    """Aplica formatação profissional ao Excel"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Cores apenas para Quimica e Quimica Industrial
    fill_quimica = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    fill_quimica_industrial = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    fill_excedente = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    font_excedente = Font(color="CC0000", bold=True)
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        
        col_widths = {
            'A': 12, 'B': 12, 'C': 18, 'D': 50, 'E': 10, 'F': 30,
            'G': 30, 'H': 12, 'I': 12, 'J': 12, 'K': 12, 'L': 12,
            'M': 12, 'N': 12, 'O': 12, 'P': 12, 'Q': 12, 'R': 12,
            'S': 12, 'T': 12, 'U': 80
        }
        
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = border
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    else:
                        if cell.column in [1, 2, 3, 5, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]:
                            cell.alignment = center_align
                        else:
                            cell.alignment = left_align
        
        # Aplicar cores apenas para Quimica (028) e Quimica Industrial (029)
        if ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                curso_cell = ws.cell(row=row, column=7)
                if curso_cell.value:
                    curso_str = str(curso_cell.value)
                    fill_color = None
                    
                    # Verifica se e exatamente Quimica (028) - nao pode ter Industrial nem Engenharia
                    if curso_str.startswith('028') or (curso_str.endswith('Química') and 'Industrial' not in curso_str and 'Engenharia' not in curso_str):
                        fill_color = fill_quimica
                    # Verifica se e exatamente Quimica Industrial (029)
                    elif curso_str.startswith('029') or 'Química Industrial' in curso_str:
                        fill_color = fill_quimica_industrial
                    # Demais cursos ficam sem cor (fundo branco)
                    
                    if fill_color:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = fill_color
                
                excedentes_cell = ws.cell(row=row, column=14)
                if excedentes_cell.value and isinstance(excedentes_cell.value, (int, float)):
                    if excedentes_cell.value > 0:
                        excedentes_cell.fill = fill_excedente
                        excedentes_cell.font = font_excedente

def gerar_excel_completo(df, periodo_str):
    """Gera Excel completo no formato do Colab"""
    if df.empty:
        return None
    
    wb = Workbook()
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    colunas_ordem = [
        'periodo', 'departamento', 'codigo_disciplina', 'nome_disciplina', 'turma', 'horarios',
        'curso_vaga', 'vagas_reg', 'vagas_vest', 'inscritos_reg', 'inscritos_vest',
        'vagas_disponiveis_reg', 'vagas_disponiveis_vest', 'excedentes', 'candidatos',
        'total_vagas', 'total_inscritos', 'total_vagas_disponiveis',
        'curso_origem_busca', 'url'
    ]
    
    for col in colunas_ordem:
        if col not in df.columns:
            df[col] = ''
    
    df = df[colunas_ordem]
    
    ws_todas = wb.create_sheet('Todas as Turmas')
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws_todas.cell(row=r_idx, column=c_idx, value=value)
    
    df_vagas_reg = df[df['vagas_disponiveis_reg'] > 0]
    if not df_vagas_reg.empty:
        ws_vagas_reg = wb.create_sheet('Com Vagas Reg')
        for r_idx, row in enumerate(dataframe_to_rows(df_vagas_reg, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_vagas_reg.cell(row=r_idx, column=c_idx, value=value)
    
    df_vagas_vest = df[df['vagas_disponiveis_vest'] > 0]
    if not df_vagas_vest.empty:
        ws_vagas_vest = wb.create_sheet('Com Vagas Vest')
        for r_idx, row in enumerate(dataframe_to_rows(df_vagas_vest, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_vagas_vest.cell(row=r_idx, column=c_idx, value=value)
    
    df_excedentes = df[df['excedentes'] > 0]
    if not df_excedentes.empty:
        ws_excedentes = wb.create_sheet('Com Excedentes')
        for r_idx, row in enumerate(dataframe_to_rows(df_excedentes, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_excedentes.cell(row=r_idx, column=c_idx, value=value)
    
    if not df.empty:
        ws_depto = wb.create_sheet('Por Departamento')
        
        grupos = df.groupby(['periodo', 'departamento'])
        
        headers = [
            'Período', 'Departamento', 'Código', 'Disciplina', 'Turma',
            'Vagas Reg', 'Vagas Vest', 'Inscritos Reg', 'Inscritos Vest',
            'Vagas Disp Reg', 'Vagas Disp Vest', 'Total Vagas', 'Total Inscritos', 'Total Vagas Disp'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_depto.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        linha_atual = 2
        
        for (periodo, departamento), grupo in grupos:
            grupo = grupo.sort_values(['codigo_disciplina', 'turma'])
            
            for _, linha in grupo.iterrows():
                dados = [
                    periodo, departamento,
                    linha['codigo_disciplina'], linha['nome_disciplina'], linha['turma'],
                    linha['vagas_reg'], linha['vagas_vest'],
                    linha['inscritos_reg'], linha['inscritos_vest'],
                    linha['vagas_disponiveis_reg'], linha['vagas_disponiveis_vest'],
                    linha['total_vagas'], linha['total_inscritos'], linha['total_vagas_disponiveis']
                ]
                
                for col, valor in enumerate(dados, 1):
                    ws_depto.cell(row=linha_atual, column=col, value=valor)
                
                linha_atual += 1
    
    ws_stats = wb.create_sheet('Estatísticas')
    
    stats_data = []
    for periodo in df['periodo'].unique():
        df_periodo = df[df['periodo'] == periodo]
        
        for curso in df_periodo['curso_vaga'].unique():
            df_curso = df_periodo[df_periodo['curso_vaga'] == curso]
            
            total_excedentes = df_curso['excedentes'].sum()
            
            stats_data.append({
                'Período': periodo,
                'Curso': curso,
                'Total Turmas': len(df_curso),
                'Turmas com Vagas Reg': len(df_curso[df_curso['vagas_disponiveis_reg'] > 0]),
                'Turmas com Vagas Vest': len(df_curso[df_curso['vagas_disponiveis_vest'] > 0]),
                'Turmas com Excedentes': len(df_curso[df_curso['excedentes'] > 0]),
                'Total Vagas Reg': df_curso['vagas_reg'].sum(),
                'Total Vagas Vest': df_curso['vagas_vest'].sum(),
                'Total Inscritos Reg': df_curso['inscritos_reg'].sum(),
                'Total Inscritos Vest': df_curso['inscritos_vest'].sum(),
                'Total Excedentes': total_excedentes,
                'Total Vagas Disp Reg': df_curso['vagas_disponiveis_reg'].sum(),
                'Total Vagas Disp Vest': df_curso['vagas_disponiveis_vest'].sum(),
                'Taxa Ocupação Reg (%)': round((df_curso['inscritos_reg'].sum() / df_curso['vagas_reg'].sum() * 100), 2) if df_curso['vagas_reg'].sum() > 0 else 0,
                'Taxa Ocupação Vest (%)': round((df_curso['inscritos_vest'].sum() / df_curso['vagas_vest'].sum() * 100), 2) if df_curso['vagas_vest'].sum() > 0 else 0
            })
    
    if stats_data:
        stats_df = pd.DataFrame(stats_data)
        for r_idx, row in enumerate(dataframe_to_rows(stats_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_stats.cell(row=r_idx, column=c_idx, value=value)
    
    aplicar_formatacao_excel(wb)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# ===== FUNÇÕES AUXILIARES =====
def formatar_periodo(periodo):
    """Formata período para exibição"""
    if '.' in periodo:
        return periodo
    elif len(periodo) == 5:
        return f"{periodo[:4]}.{periodo[4]}"
    return periodo

def validar_periodo(periodo):
    """Valida formato do período"""
    if '.' in periodo:
        partes = periodo.split('.')
        if len(partes) == 2 and partes[0].isdigit() and partes[1].isdigit():
            ano = int(partes[0])
            semestre = int(partes[1])
            if 2000 <= ano <= 2100 and semestre in [1, 2]:
                return True
    return False

def validar_departamento(depto):
    """Valida formato do departamento"""
    if depto == 'TODOS' or depto == '':
        return True
    if len(depto) == 3 and depto.isalpha():
        return True
    return False

def validar_codigo_disciplina(codigo):
    """Valida formato do código de disciplina (3 letras + 5 números)"""
    if not codigo:
        return False
    codigo = codigo.strip().upper()
    # Padrão: 3 letras seguidas de 5 números (ex: GQI00061)
    padrao = r'^[A-Z]{3}\d{5}$'
    return bool(re.match(padrao, codigo))

def criar_visualizacoes(df):
    """Cria visualizações gráficas dos dados"""
    if df.empty:
        st.info("📭 Nenhum dado disponível para visualização")
        return
    
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Visão Geral", "📈 Distribuição", "🏫 Análise Detalhada", "⚠️ Excedentes"])
    
    with tab1:
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_turmas = len(df)
            st.metric("Total de Turmas", total_turmas)
        
        with col2:
            turmas_com_vagas = len(df[df['total_vagas_disponiveis'] > 0])
            st.metric("Turmas com Vagas", turmas_com_vagas)
        
        with col3:
            total_vagas_disp = df['total_vagas_disponiveis'].sum()
            st.metric("Vagas Disponíveis", total_vagas_disp)
        
        with col4:
            total_excedentes = df['excedentes'].sum()
            st.metric("Total de Excedentes", total_excedentes, delta=None)
        
        st.subheader("📊 Vagas Disponíveis por Curso")
        
        vagas_curso = df.groupby('curso_vaga').agg({
            'vagas_disponiveis_reg': 'sum',
            'vagas_disponiveis_vest': 'sum'
        }).reset_index()
        
        if not vagas_curso.empty:
            fig = go.Figure()
            
            fig.add_trace(go.Bar(
                name='Vagas Regulares',
                x=vagas_curso['curso_vaga'],
                y=vagas_curso['vagas_disponiveis_reg'],
                marker_color='#1e3a5f'
            ))
            
            fig.add_trace(go.Bar(
                name='Vagas Vestibular',
                x=vagas_curso['curso_vaga'],
                y=vagas_curso['vagas_disponiveis_vest'],
                marker_color='#4a90e2'
            ))
            
            fig.update_layout(
                barmode='stack',
                height=400,
                title="Vagas Disponíveis por Tipo e Curso",
                xaxis_title="Curso",
                yaxis_title="Vagas Disponíveis",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            
            st.plotly_chart(fig, use_container_width=True)
    
    with tab2:
        st.subheader("🏫 Distribuição por Departamento")
        
        depto_dist = df.groupby('departamento').agg({
            'codigo_disciplina': 'count',
            'total_vagas_disponiveis': 'sum',
            'excedentes': 'sum'
        }).reset_index()
        depto_dist.columns = ['Departamento', 'Número de Turmas', 'Vagas Disponíveis', 'Excedentes']
        
        if not depto_dist.empty:
            col1, col2 = st.columns(2)
            
            with col1:
                fig = px.treemap(
                    depto_dist,
                    path=['Departamento'],
                    values='Vagas Disponíveis',
                    color='Excedentes',
                    color_continuous_scale='Reds',
                    title='Vagas Disponíveis por Departamento'
                )
                fig.update_layout(height=500)
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                st.write("**Ranking de Departamentos:**")
                depto_ranking = depto_dist.sort_values('Excedentes', ascending=False)
                st.dataframe(
                    depto_ranking,
                    column_config={
                        "Departamento": st.column_config.TextColumn("Depto"),
                        "Número de Turmas": st.column_config.NumberColumn("Turmas"),
                        "Vagas Disponíveis": st.column_config.NumberColumn("Vagas Disp."),
                        "Excedentes": st.column_config.NumberColumn("Excedentes")
                    },
                    hide_index=True,
                    use_container_width=True
                )
    
    with tab3:
        st.subheader("📋 Análise Detalhada por Disciplina")
        
        col_filt1, col_filt2 = st.columns(2)
        
        with col_filt1:
            curso_analise = st.selectbox(
                "Selecione o curso para análise:",
                options=['Todos'] + list(df['curso_vaga'].unique()),
                key="analise_curso"
            )
        
        with col_filt2:
            ordenacao = st.selectbox(
                "Ordenar por:",
                options=['Mais vagas disponíveis', 'Mais inscritos', 'Mais excedentes', 'Código da disciplina'],
                key="analise_ordenacao"
            )
        
        if curso_analise != 'Todos':
            df_analise = df[df['curso_vaga'] == curso_analise].copy()
        else:
            df_analise = df.copy()
        
        if ordenacao == 'Mais vagas disponíveis':
            df_analise = df_analise.sort_values('total_vagas_disponiveis', ascending=False)
        elif ordenacao == 'Mais inscritos':
            df_analise = df_analise.sort_values('total_inscritos', ascending=False)
        elif ordenacao == 'Mais excedentes':
            df_analise = df_analise.sort_values('excedentes', ascending=False)
        else:
            df_analise = df_analise.sort_values(['codigo_disciplina', 'turma'])
        
        st.dataframe(
            df_analise[[
                'codigo_disciplina', 'nome_disciplina', 'turma', 'horarios',
                'vagas_reg', 'inscritos_reg', 'vagas_disponiveis_reg',
                'vagas_vest', 'inscritos_vest', 'vagas_disponiveis_vest',
                'excedentes', 'candidatos', 'total_vagas_disponiveis'
            ]].head(20),
            column_config={
                "codigo_disciplina": "Código",
                "nome_disciplina": "Disciplina",
                "turma": "Turma",
                "horarios": "Horários",
                "vagas_reg": "Vagas Reg",
                "inscritos_reg": "Inscritos Reg",
                "vagas_disponiveis_reg": "Disp. Reg",
                "vagas_vest": "Vagas Vest",
                "inscritos_vest": "Inscritos Vest",
                "vagas_disponiveis_vest": "Disp. Vest",
                "excedentes": st.column_config.NumberColumn("Excedentes"),
                "candidatos": "Candidatos",
                "total_vagas_disponiveis": "Total Disp."
            },
            hide_index=True,
            use_container_width=True
        )
    
    with tab4:
        st.subheader("⚠️ Análise de Excedentes")
        
        df_excedentes = df[df['excedentes'] > 0].copy()
        
        if not df_excedentes.empty:
            st.warning(f"⚠️ **Atenção:** Foram encontradas {len(df_excedentes)} turmas com excedentes!")
            
            col_ex1, col_ex2, col_ex3 = st.columns(3)
            
            with col_ex1:
                total_excedentes = df_excedentes['excedentes'].sum()
                st.metric("Total de Excedentes", total_excedentes)
            
            with col_ex2:
                cursos_com_excedentes = len(df_excedentes['curso_vaga'].unique())
                st.metric("Cursos com Excedentes", cursos_com_excedentes)
            
            with col_ex3:
                maior_excedente = df_excedentes['excedentes'].max()
                st.metric("Maior Excedente", maior_excedente)
            
            st.subheader("📋 Turmas com Excedentes")
            
            df_excedentes_ordenado = df_excedentes.sort_values('excedentes', ascending=False)
            
            st.dataframe(
                df_excedentes_ordenado[[
                    'codigo_disciplina', 'nome_disciplina', 'turma', 'curso_vaga',
                    'vagas_reg', 'candidatos', 'excedentes', 'inscritos_reg'
                ]],
                column_config={
                    "codigo_disciplina": "Código",
                    "nome_disciplina": "Disciplina",
                    "turma": "Turma",
                    "curso_vaga": "Curso",
                    "vagas_reg": "Vagas Reg",
                    "candidatos": "Candidatos",
                    "excedentes": st.column_config.NumberColumn("Excedentes"),
                    "inscritos_reg": "Inscritos Reg"
                },
                hide_index=True,
                use_container_width=True
            )
            
            st.subheader("📊 Excedentes por Curso")
            
            excedentes_curso = df_excedentes.groupby('curso_vaga').agg({
                'excedentes': 'sum',
                'codigo_disciplina': 'count'
            }).reset_index()
            excedentes_curso.columns = ['Curso', 'Total Excedentes', 'Número de Turmas']
            excedentes_curso = excedentes_curso.sort_values('Total Excedentes', ascending=False)
            
            col_exc1, col_exc2 = st.columns(2)
            
            with col_exc1:
                fig = px.bar(
                    excedentes_curso,
                    x='Curso',
                    y='Total Excedentes',
                    color='Total Excedentes',
                    color_continuous_scale='Reds',
                    title='Total de Excedentes por Curso'
                )
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)
            
            with col_exc2:
                st.dataframe(
                    excedentes_curso,
                    column_config={
                        "Curso": st.column_config.TextColumn("Curso"),
                        "Total Excedentes": st.column_config.NumberColumn("Excedentes"),
                        "Número de Turmas": st.column_config.NumberColumn("Turmas")
                    },
                    hide_index=True,
                    use_container_width=True
                )
        else:
            st.success("✅ Nenhuma turma com excedentes encontrada!")

# ===== INTERFACE PRINCIPAL =====
st.markdown("""
<div class="main-header-container">
    <h1 class="main-header">Consultor de Vagas UFF</h1>
    <p class="sub-header">Instituto de Quimica - Sistema de Consulta Detalhada de Turmas</p>
    <p class="developer-credit">Desenvolvido por <span class="developer-name">Tadeu L. Araujo</span> (GGQ)</p>
</div>
""", unsafe_allow_html=True)

# Sidebar com filtros
with st.sidebar:
    st.markdown("""
    <div style="text-align: center; padding: 1rem 0; margin-bottom: 1rem;">
        <h2 style="color: #1e3a5f; margin: 0; font-size: 1.5rem; font-weight: 700;">Configuracoes</h2>
        <div style="height: 3px; background: linear-gradient(90deg, transparent, #3b82f6, transparent); margin-top: 0.5rem;"></div>
    </div>
    """, unsafe_allow_html=True)
    
    # === SEÇÃO: PERÍODO ===
    st.subheader("📅 Período Letivo")
    
    periodo_input = st.text_input(
        "Digite o período (ex: 2025.2):",
        value="2025.2",
        help="Formato: AAAA.S (ex: 2025.2 para 2025 semestre 2)",
        key="periodo_input"
    )
    
    periodos_formatados = []
    
    if periodo_input:
        if validar_periodo(periodo_input):
            periodos_formatados = [periodo_input.replace('.', '')]
            st.success(f"✅ {periodo_input}")
        else:
            st.error("❌ Formato inválido")
            periodos_formatados = []
    
    adicionar_periodo = st.checkbox("Adicionar outro período", key="adicionar_periodo")
    if adicionar_periodo:
        periodo2 = st.text_input("Segundo período:", value="2025.1", key="periodo2")
        if periodo2 and validar_periodo(periodo2):
            periodos_formatados.append(periodo2.replace('.', ''))
    
    st.markdown("---")
    
    # === SEÇÃO: CURSOS ===
    st.subheader("🎓 Cursos")
    
    cursos_selecionados = st.multiselect(
        "Selecione os cursos:",
        options=['Química', 'Química Industrial', 'Engenharia Química', 'Farmácia'],
        default=['Química', 'Química Industrial'],
        key="cursos_selecionados"
    )
    
    st.markdown("---")
    
    # === SEÇÃO: DISCIPLINA ESPECÍFICA ===
    st.subheader("📚 Disciplina Específica")
    
    codigo_disciplina_input = st.text_input(
        "Código da disciplina (opcional):",
        value="",
        max_chars=8,
        help="Digite o código completo (ex: GQI00061) para consultar uma disciplina específica",
        key="codigo_disciplina_input",
        placeholder="Ex: GQI00061"
    )
    
    codigo_disciplina_valido = None
    if codigo_disciplina_input:
        if validar_codigo_disciplina(codigo_disciplina_input):
            codigo_disciplina_valido = codigo_disciplina_input.strip().upper()
            st.success(f"✅ Disciplina: {codigo_disciplina_valido}")
        else:
            st.error("❌ Formato: 3 letras + 5 números")
    
    st.markdown("---")
    
    # === SEÇÃO: DEPARTAMENTOS ===
    st.subheader("🏫 Departamentos")
    
    modo_departamento = st.radio(
        "Modo de seleção:",
        options=['Lista pré-definida', 'Digitar código'],
        key="modo_departamento",
        horizontal=True
    )
    
    departamentos_selecionados = []
    
    if modo_departamento == 'Lista pré-definida':
        # Lista atualizada de departamentos
        departamentos_opcoes = [
            'TODOS', 'GGQ', 'GQI', 'GQA', 'GQO', 'GFQ', 'GEO', 'GMA', 
            'GFI', 'SSE', 'TEQ', 'TEP', 'TDT', 'SFP', 'GLC', 'GGM', 'MTC', 'GCM'
        ]
        
        departamentos_selecionados = st.multiselect(
            "Selecione departamentos:",
            options=departamentos_opcoes,
            default=['TODOS'],
            key="departamentos_lista"
        )
    else:
        depto_input = st.text_input(
            "Código do departamento (3 letras):",
            value="GQI",
            max_chars=3,
            help="Ex: GQI, GGQ, TEQ, etc.",
            key="depto_input"
        )
        
        if depto_input:
            depto_input = depto_input.strip().upper()
            if validar_departamento(depto_input):
                departamentos_selecionados = [depto_input]
                st.success(f"✅ {depto_input}")
            else:
                st.error("❌ Use 3 letras ou 'TODOS'")
                departamentos_selecionados = []
        else:
            departamentos_selecionados = ['TODOS']
    
    st.markdown("---")
    
    # === CONFIGURAÇÕES AVANÇADAS ===
    with st.expander("⚙️ Configurações Avançadas"):
        apenas_cursos_quimica_checkbox = st.checkbox(
            "Mostrar apenas cursos selecionados", 
            value=st.session_state.apenas_cursos_quimica,
            help="Filtrar para mostrar apenas vagas dos cursos selecionados acima",
            key="apenas_cursos_quimica_checkbox"
        )
        
        st.session_state.apenas_cursos_quimica = apenas_cursos_quimica_checkbox
        
        mostrar_outros_cursos_checkbox = st.checkbox(
            "Mostrar também vagas de outros cursos", 
            value=st.session_state.mostrar_outros_cursos,
            help="Mostrar vagas de todos os cursos, não apenas dos selecionados",
            key="mostrar_outros_cursos_checkbox"
        )
        
        st.session_state.mostrar_outros_cursos = mostrar_outros_cursos_checkbox
    
    st.markdown("---")
    
    # === BOTÕES DE AÇÃO ===
    col1, col2 = st.columns(2)
    
    with col1:
        btn_consultar = st.button(
            "🔍 Consultar", 
            type="primary", 
            use_container_width=True,
            disabled=not periodos_formatados or not cursos_selecionados,
            key="btn_consultar"
        )
    
    with col2:
        btn_limpar = st.button("🔄 Limpar", use_container_width=True, key="btn_limpar")
        if btn_limpar:
            st.session_state.processando = False
            st.session_state.resultado_disponivel = False
            st.session_state.dados_turmas = None
            st.rerun()
    
    st.markdown("---")
    
    st.markdown("""
    <div style="background: linear-gradient(145deg, #eff6ff, #dbeafe); border-radius: 10px; padding: 1rem; border-left: 3px solid #3b82f6;">
        <p style="color: #1e293b; font-size: 0.9rem; margin: 0;">
            <strong style="color: #1e3a5f;">Dicas:</strong><br>
            - A consulta pode levar alguns minutos<br>
            - Para disciplina especifica, use o codigo completo (ex: GQI00061)<br>
            - Os dados sao extraidos em tempo real
        </p>
    </div>
    """, unsafe_allow_html=True)

# Área principal - Processamento
if btn_consultar and periodos_formatados and cursos_selecionados:
    st.session_state.processando = True
    st.session_state.resultado_disponivel = False
    
    with st.spinner("🔄 Inicializando consulta..."):
        try:
            consultor = ConsultorQuadroHorariosUFFDetalhado(
                apenas_cursos_quimica=st.session_state.apenas_cursos_quimica,
                mostrar_outros_cursos=st.session_state.mostrar_outros_cursos,
                cursos_selecionados=cursos_selecionados
            )
            
            deptos_consulta = []
            for depto in departamentos_selecionados:
                if depto == 'TODOS':
                    deptos_consulta.append(None)
                else:
                    deptos_consulta.append(depto)
            
            if not deptos_consulta:
                deptos_consulta = [None]
            
            # Mostrar configuração da consulta
            config_msg = f"""
            **🎯 Consulta Configurada:**
            - 📅 Períodos: {', '.join([formatar_periodo(p) for p in periodos_formatados])}
            - 🎓 Cursos: {', '.join(cursos_selecionados)}
            - 🏫 Departamentos: {', '.join([d if d else 'Todos' for d in departamentos_selecionados])}
            """
            if codigo_disciplina_valido:
                config_msg += f"\n- 📚 Disciplina específica: {codigo_disciplina_valido}"
            
            st.info(config_msg)
            
            dados = consultor.consultar_vagas_completas(
                periodos=periodos_formatados,
                cursos=cursos_selecionados,
                departamentos=deptos_consulta,
                codigo_disciplina=codigo_disciplina_valido
            )
            
            if dados:
                df_resultado = pd.DataFrame(dados)
                
                df_resultado['excedentes'] = pd.to_numeric(df_resultado['excedentes'], errors='coerce').fillna(0)
                df_resultado['candidatos'] = pd.to_numeric(df_resultado['candidatos'], errors='coerce').fillna(0)
                df_resultado['vagas_reg'] = pd.to_numeric(df_resultado['vagas_reg'], errors='coerce').fillna(0)
                
                for idx, row in df_resultado.iterrows():
                    if row['excedentes'] == 0 and row['candidatos'] > 0 and row['vagas_reg'] > 0:
                        if row['candidatos'] > row['vagas_reg']:
                            df_resultado.at[idx, 'excedentes'] = row['candidatos'] - row['vagas_reg']
                
                st.session_state.dados_turmas = df_resultado
                st.session_state.resultado_disponivel = True
                st.session_state.processando = False
                
                st.success(f"✅ Consulta concluída! {len(dados)} registros coletados")
                
                col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                with col_stat1:
                    st.metric("Cursos", len(df_resultado['curso_vaga'].unique()))
                with col_stat2:
                    st.metric("Departamentos", len(df_resultado['departamento'].unique()))
                with col_stat3:
                    st.metric("Turmas com vagas", len(df_resultado[df_resultado['total_vagas_disponiveis'] > 0]))
                with col_stat4:
                    st.metric("Com excedentes", len(df_resultado[df_resultado['excedentes'] > 0]))
                
                st.rerun()
            else:
                st.error("❌ Nenhuma turma encontrada com os filtros selecionados.")
                st.session_state.processando = False
        
        except Exception as e:
            st.error(f"❌ Erro durante a consulta: {str(e)}")
            st.exception(e)
            st.session_state.processando = False

# Area principal - Resultados
if st.session_state.resultado_disponivel and st.session_state.dados_turmas is not None:
    df = st.session_state.dados_turmas
    
    st.markdown('<div class="custom-divider"></div>', unsafe_allow_html=True)
    st.markdown('<p class="section-header">Resultados da Consulta</p>', unsafe_allow_html=True)
    
    if periodos_formatados:
        periodo_formatado = formatar_periodo(periodos_formatados[0])
    else:
        periodo_formatado = "N/A"
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Período", periodo_formatado)
    
    with col2:
        st.metric("Total de Turmas", len(df))
    
    with col3:
        turmas_vagas = len(df[df['total_vagas_disponiveis'] > 0])
        st.metric("Turmas com Vagas", turmas_vagas)
    
    with col4:
        total_excedentes = df['excedentes'].sum()
        st.metric("Total de Excedentes", total_excedentes, delta=None)
    
    # Visualizações
    criar_visualizacoes(df)
    
    # Exportacao - APENAS EXCEL
    st.markdown('<div class="custom-divider"></div>', unsafe_allow_html=True)
    st.markdown('<p class="section-header">Exportar Resultados</p>', unsafe_allow_html=True)
    
    col_exp1, col_exp2, col_exp3 = st.columns([1, 2, 1])
    
    with col_exp2:
        excel_buffer = gerar_excel_completo(df, periodo_formatado)
        if excel_buffer:
            st.download_button(
                label="📊 Baixar Excel Completo",
                data=excel_buffer,
                file_name=f"vagas_uff_detalhado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_download_excel"
            )
        else:
            st.warning("⚠️ Nenhum dado para exportar")
    
    # Tabela interativa completa
    st.markdown('<div class="custom-divider"></div>', unsafe_allow_html=True)
    st.markdown('<p class="section-header">Tabela Completa de Dados</p>', unsafe_allow_html=True)
    
    col_filt1, col_filt2, col_filt3 = st.columns(3)
    
    with col_filt1:
        filtro_curso = st.selectbox(
            "Filtrar por curso:",
            options=['Todos'] + list(df['curso_vaga'].unique()),
            key="filtro_curso_tabela"
        )
    
    with col_filt2:
        filtro_depto = st.selectbox(
            "Filtrar por departamento:",
            options=['Todos'] + [d for d in df['departamento'].unique() if pd.notna(d)],
            key="filtro_depto_tabela"
        )
    
    with col_filt3:
        filtro_vagas = st.selectbox(
            "Filtrar por vagas:",
            options=['Todas', 'Com vagas disponíveis', 'Sem vagas', 'Com excedentes'],
            key="filtro_vagas_tabela"
        )
    
    df_filtrado = df.copy()
    
    if filtro_curso != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['curso_vaga'] == filtro_curso]
    
    if filtro_depto != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['departamento'] == filtro_depto]
    
    if filtro_vagas == 'Com vagas disponíveis':
        df_filtrado = df_filtrado[df_filtrado['total_vagas_disponiveis'] > 0]
    elif filtro_vagas == 'Sem vagas':
        df_filtrado = df_filtrado[df_filtrado['total_vagas_disponiveis'] == 0]
    elif filtro_vagas == 'Com excedentes':
        df_filtrado = df_filtrado[df_filtrado['excedentes'] > 0]
    
    st.dataframe(
        df_filtrado[[
            'periodo', 'departamento', 'codigo_disciplina', 'nome_disciplina', 
            'turma', 'curso_vaga', 'vagas_reg', 'inscritos_reg', 'vagas_disponiveis_reg',
            'vagas_vest', 'inscritos_vest', 'vagas_disponiveis_vest', 'excedentes', 'candidatos', 'total_vagas_disponiveis'
        ]],
        column_config={
            "periodo": "Período",
            "departamento": "Depto",
            "codigo_disciplina": "Código",
            "nome_disciplina": "Disciplina",
            "turma": "Turma",
            "curso_vaga": "Curso",
            "vagas_reg": "Vagas Reg",
            "inscritos_reg": "Inscritos Reg",
            "vagas_disponiveis_reg": "Disp. Reg",
            "vagas_vest": "Vagas Vest",
            "inscritos_vest": "Inscritos Vest",
            "vagas_disponiveis_vest": "Disp. Vest",
            "excedentes": st.column_config.NumberColumn("Excedentes"),
            "candidatos": "Candidatos",
            "total_vagas_disponiveis": "Total Disp."
        },
        hide_index=True,
        use_container_width=True,
        height=400
    )
    
    st.info(f"Mostrando {len(df_filtrado)} de {len(df)} registros")

# Pagina inicial
elif not st.session_state.processando:
    st.markdown('<div class="custom-divider"></div>', unsafe_allow_html=True)
    
    # Cards de introducao
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="info-card">
            <h3 style="color: #1e40af; margin-bottom: 0.5rem;">Como Usar</h3>
            <ol style="color: #334155; margin: 0; padding-left: 1.2rem;">
                <li>Digite o periodo (ex: 2025.2)</li>
                <li>Selecione os cursos</li>
                <li>Escolha departamentos</li>
                <li>Clique em Consultar</li>
                <li>Exporte em Excel</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="success-card">
            <h3 style="color: #065f46; margin-bottom: 0.5rem;">Disciplina Especifica</h3>
            <p style="color: #334155; margin: 0;">
                <strong>Formato:</strong> 3 letras + 5 numeros<br>
                <strong>Exemplos:</strong><br>
                GQI00061, TEQ00042<br>
                GMA00159, GFI00025
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="warning-card">
            <h3 style="color: #92400e; margin-bottom: 0.5rem;">Departamentos</h3>
            <p style="color: #334155; margin: 0; font-size: 0.9rem;">
                <strong>Disponiveis:</strong> TODOS (padrao)<br>
                GGQ, GQI, GQA, GQO, GFQ<br>
                GEO, GMA, GFI, SSE, TEQ<br>
                TEP, TDT, SFP, GLC, GGM<br>
                MTC, GCM
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Exemplo de dados em expander estilizado
    with st.expander("Ver exemplo de dados coletados"):
        st.markdown("""
        | Campo | Descricao | Exemplo |
        |-------|-----------|---------|
        | **periodo** | Periodo letivo | 20252 |
        | **departamento** | Codigo do departamento | GQI |
        | **codigo_disciplina** | Codigo da disciplina | GQI00061 |
        | **nome_disciplina** | Nome da disciplina | Quimica Geral |
        | **turma** | Identificacao da turma | A01 |
        | **curso_vaga** | Curso da vaga | 028 - Quimica |
        | **vagas_reg** | Vagas regulares | 40 |
        | **inscritos_reg** | Inscritos regulares | 35 |
        | **vagas_disponiveis_reg** | Vagas disp. regulares | 5 |
        | **excedentes** | Excedentes calculados | 0 |
        """)

# Rodapé
st.markdown(f"""
<div class="footer-container">
    <p class="footer-text">
        <strong>Consultor de Vagas UFF</strong> - Instituto de Quimica<br>
        Desenvolvido por <span class="footer-highlight">Tadeu L. Araujo (GGQ)</span><br>
        Versao: {datetime.now().strftime('%d/%m/%Y')}
    </p>
</div>
""", unsafe_allow_html=True)

if st.session_state.processando:
    st.warning("⏳ Processamento em andamento... Por favor, aguarde.")
